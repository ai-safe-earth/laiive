"""Batch upload: CSV/XLSX → EventDrafts (02-arch §4).

The server parses rows into drafts; the frontend shows one form at a time
("event i of N") and each approval triggers one write via /batch/validate-event.
"""

import csv
import io

from laiive_shared import EventDraft, missing_required
from loguru import logger

# Flexible header → draft-field mapping (lowercased, stripped).
HEADER_ALIASES: dict[str, str] = {
    "name": "name",
    "event": "name",
    "event_name": "name",
    "title": "name",
    "artist": "artists",
    "artists": "artists",
    "lineup": "artists",
    "date": "start_at",
    "datetime": "start_at",
    "date_time": "start_at",
    "start": "start_at",
    "start_at": "start_at",
    "when": "start_at",
    "venue": "venue",
    "place": "venue",
    "address": "address",
    "city": "city",
    "town": "city",
    "venue_type": "venue_type",
    "type": "venue_type",
    "price": "price_min",
    "price_min": "price_min",
    "price_max": "price_max",
    "currency": "price_currency",
    "price_currency": "price_currency",
    "description": "description",
    "about": "description",
    "details": "description",
    "genre": "genre",
    "style": "genre",
    "tickets": "ticket_url",
    "ticket_url": "ticket_url",
    "url": "ticket_url",
    "link": "ticket_url",
}


def parse_batch(file_bytes: bytes, filename: str) -> list[EventDraft]:
    """Parse an uploaded spreadsheet into drafts, one per data row."""
    name = filename.lower()
    if name.endswith(".csv"):
        rows = _read_csv(file_bytes)
    elif name.endswith((".xlsx", ".xls")):
        rows = _read_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename} (use .csv or .xlsx)")
    return [draft for row in rows if (draft := _row_to_draft(row)) is not None]


def _read_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return list(csv.DictReader(io.StringIO(text), dialect=dialect))


def _read_xlsx(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    return [
        {h: cell for h, cell in zip(headers, row) if h}
        for row in rows
        if any(cell is not None for cell in row)
    ]


def _row_to_draft(row: dict) -> EventDraft | None:
    data: dict = {}
    for header, value in row.items():
        if value is None or str(value).strip() == "":
            continue
        field = HEADER_ALIASES.get(str(header).strip().lower().replace(" ", "_"))
        if field is None:
            continue
        data[field] = value

    if not data:
        return None

    if "artists" in data:
        raw = str(data["artists"])
        separator = ";" if ";" in raw else ","
        data["artists"] = [a.strip() for a in raw.split(separator) if a.strip()]
    for price_key in ("price_min", "price_max"):
        if price_key in data:
            try:
                data[price_key] = float(str(data[price_key]).replace(",", "."))
            except ValueError:
                logger.warning(f"Unparseable {price_key}: {data[price_key]!r}")
                del data[price_key]
    for key in (
        "name",
        "start_at",
        "venue",
        "address",
        "city",
        "venue_type",
        "price_currency",
        "description",
        "genre",
        "ticket_url",
    ):
        if key in data:
            data[key] = str(data[key]).strip()

    try:
        return EventDraft(**data)
    except Exception as e:
        logger.warning(f"Row could not become a draft: {e}")
        return None


def drafts_with_missing(drafts: list[EventDraft]) -> list[dict]:
    """Serializable [draft + its missing required fields] for the parse response."""
    return [
        {"draft": d.model_dump(exclude_none=True), "missing": missing_required(d)}
        for d in drafts
    ]
